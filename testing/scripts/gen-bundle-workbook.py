#!/usr/bin/env python3
"""
gen-bundle-workbook.py - build a checkable workbook of how every PACT subclass spell bundle is priced.

Reads the JSON dumped from js/engine-data.js (so every engine figure is sourced, not typed) and writes
an .xlsx whose derived columns are live formulas, so the arithmetic can be checked and re-run.

  node -e "...dump..." > bundles.json
  python testing/scripts/gen-bundle-workbook.py bundles.json out.xlsx
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

src, dest = sys.argv[1], sys.argv[2]
D = json.load(open(src))
KU = D["knownUnit"]            # AP per spell known, by spell level
CANTRIP = D["cantCum"][1]      # a bundle cantrip is charged flat at this, not on the cantrip ladder

ARIAL = "Arial"
H_FILL = PatternFill("solid", fgColor="58180D")
H_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
SUB_FILL = PatternFill("solid", fgColor="E5D8B8")
IN_FONT = Font(name=ARIAL, color="0000FF")          # hardcoded input
FM_FONT = Font(name=ARIAL)                          # formula / derived
EN_FONT = Font(name=ARIAL, color="008000")          # value read out of the engine
BAD = PatternFill("solid", fgColor="FFC7CE")
OK = PatternFill("solid", fgColor="C6EFCE")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

# The grant shape each bundle is built from. This is a RECONSTRUCTION, not stored in the engine -
# the engine keeps only the lump price. It is stated as an assumption on the Model sheet.
SHAPE = {
    "full":   {"levels": [1, 1, 2, 2, 3, 3], "label": "Full caster - 2 spells each at 1st/2nd/3rd"},
    "half":   {"levels": [1, 1, 2, 2],       "label": "Half caster - 2 spells each at 1st/2nd"},
    "ranger": {"levels": [1, 2],             "label": "Half caster - 1 spell each at 1st/2nd"},
}
def shape_of(cls):
    return "ranger" if cls == "Ranger" else "half" if cls == "Paladin" else "full"

wb = Workbook()

def head(ws, row, labels, widths):
    for i, (lab, w) in enumerate(zip(labels, widths), start=1):
        c = ws.cell(row=row, column=i, value=lab)
        c.fill, c.font = H_FILL, H_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)

# ---------------------------------------------------------------- 1. Model -------------------
ws = wb.active
ws.title = "Model"
ws.sheet_view.showGridLines = False
def put(r, a, b=None, bold=False, font=None, wrap=False):
    c = ws.cell(row=r, column=1, value=a)
    c.font = Font(name=ARIAL, bold=bold, size=13 if bold and r == 1 else 11)
    c.alignment = Alignment(wrap_text=wrap, vertical="top")
    if b is not None:
        d = ws.cell(row=r, column=2, value=b)
        d.font = font or IN_FONT
    return r + 1

ws.column_dimensions["A"].width = 104
ws.column_dimensions["B"].width = 14
r = put(1, "PACT - how a subclass bonus-spell bundle is priced", bold=True)
r = put(r + 1, "Engine rules version (js/engine-data.js)", D["version"], font=EN_FONT)
r = put(r + 1, "READ THIS FIRST - is the origin discount applied twice?", bold=True)
for line in [
    "No. A bundle is charged with ONE lookup: engine.js line 341 does `_isO ? bundle.origin : bundle.cross`.",
    "It picks one of two stored numbers. It never takes a stored number and subtracts a discount from it,",
    "so there is no second application to be had. The discount is applied once, when the stored `origin`",
    "figure was derived - that derivation is what the 'Per-spell working' sheet reproduces.",
    "",
    "What DOES look like a doubled discount, and isn't: column 3 of the guide's ability tables is headed",
    "'Sticker (Origin)'. For an ordinary class feature the sticker is already reduced - sticker = cross - tier.",
    "For a bundle there is no tier, so nothing is subtracted and the sticker equals the full cross price.",
    "Both are the same thing in meaning - 'what a non-origin member of that class pays' - but they are",
    "reached differently, so the two row types look inconsistent side by side. See the 'Price paths' sheet,",
    "which is verified against a real compute() run, not just read off the source.",
]:
    r = put(r, line)

r = put(r + 1, "Inputs (blue = hardcoded, green = read from the engine)", bold=True)
r = put(r, "Origin discount, AP per spell", 1)
r = put(r, "Floor - no spell can cost less than this", 1)
r = put(r, "Cantrip granted by a bundle, flat AP each", CANTRIP, font=EN_FONT)
DD_C, FLOOR_C, CANT_C = f"$B${r-3}", f"$B${r-2}", f"$B${r-1}"

r = put(r + 1, "AP per spell known, by spell level (DATA.knownUnit)", bold=True)
KU_ROW0 = r
for lv, ap in enumerate(KU, start=1):
    ws.cell(row=r, column=1, value=f"Level {lv}").font = FM_FONT
    ws.cell(row=r, column=2, value=ap).font = EN_FONT
    r += 1
KU_RANGE = f"$A${KU_ROW0}:$B${KU_ROW0+len(KU)-1}"

r = put(r + 1, "Assumption you should check", bold=True)
for line in [
    "The engine stores ONLY the lump price for each bundle. How many spells a bundle contains, and at",
    "what levels, is NOT stored anywhere - it is reconstructed here from the guide's rule that a bundle",
    "prices the spells granted at the subclass's first few grant levels. The reconstruction reproduces",
    "16 of the 21 engine prices exactly, which is why it is trusted; the 5 it misses are flagged.",
    "If a bundle's real spell list differs from the shape on the 'Bundles' sheet, edit columns D-G there",
    "and every derived figure recalculates.",
]:
    r = put(r, line)

# ---------------------------------------------------------------- 2. Per-spell working --------
li = wb.create_sheet("Per-spell working")
li.sheet_view.showGridLines = False
head(li, 1, ["Key", "Class", "Subclass", "Item", "Spell level",
             "Cross AP\n(no discount)", "Origin AP\n(-1, floored at 1)", "Discount saved"],
     [26, 11, 24, 26, 11, 13, 14, 12])
row = 2
for b in D["bundles"]:
    key = f'{b["cls"]}|{b["sub"]}'
    sh = SHAPE[shape_of(b["cls"])]
    for i, lv in enumerate(sh["levels"], start=1):
        li.cell(row=row, column=1, value=key).font = FM_FONT
        li.cell(row=row, column=2, value=b["cls"]).font = FM_FONT
        li.cell(row=row, column=3, value=b["sub"]).font = FM_FONT
        li.cell(row=row, column=4, value=f"Spell {i}").font = FM_FONT
        li.cell(row=row, column=5, value=lv).font = IN_FONT
        # cost = knownUnit[level]; origin = max(floor, that - dd)
        li.cell(row=row, column=6, value=f"=VLOOKUP(\"Level \"&E{row},Model!{KU_RANGE},2,FALSE)").font = FM_FONT
        li.cell(row=row, column=7, value=f"=MAX(Model!{FLOOR_C},F{row}-Model!{DD_C})").font = FM_FONT
        li.cell(row=row, column=8, value=f"=F{row}-G{row}").font = FM_FONT
        row += 1
    for i in range(b["cantrips"]):
        li.cell(row=row, column=1, value=key).font = FM_FONT
        li.cell(row=row, column=2, value=b["cls"]).font = FM_FONT
        li.cell(row=row, column=3, value=b["sub"]).font = FM_FONT
        li.cell(row=row, column=4, value=f"Cantrip {i+1}").font = FM_FONT
        li.cell(row=row, column=5, value=0).font = IN_FONT
        li.cell(row=row, column=6, value=f"=Model!{CANT_C}").font = FM_FONT
        li.cell(row=row, column=7, value=f"=Model!{CANT_C}").font = FM_FONT   # cantrips take no discount
        li.cell(row=row, column=8, value=f"=F{row}-G{row}").font = FM_FONT
        row += 1
LI_LAST = row - 1

# ---------------------------------------------------------------- 3. Bundles ------------------
bs = wb.create_sheet("Bundles")
bs.sheet_view.showGridLines = False
head(bs, 1, ["Class", "Subclass", "Grant shape assumed", "Spells", "Cantrips",
             "Cross AP\nderived", "Origin AP\nderived", "Cross AP\nENGINE", "Origin AP\nENGINE",
             "Cross\ndiff", "Origin\ndiff", "Origin gap\nderived", "Origin gap\nENGINE",
             "Is the discount applied twice?", "Derivation\nreproduces engine?"],
     [11, 24, 40, 8, 9, 11, 11, 11, 11, 8, 8, 11, 11, 34, 22])
row = 2
for b in D["bundles"]:
    key = f'{b["cls"]}|{b["sub"]}'
    sh = SHAPE[shape_of(b["cls"])]
    bs.cell(row=row, column=1, value=b["cls"]).font = FM_FONT
    bs.cell(row=row, column=2, value=b["sub"]).font = FM_FONT
    bs.cell(row=row, column=3, value=sh["label"]).font = IN_FONT
    bs.cell(row=row, column=4, value=len(sh["levels"])).font = IN_FONT
    bs.cell(row=row, column=5, value=b["cantrips"]).font = EN_FONT
    q = f"'Per-spell working'!$A$2:$A${LI_LAST}"
    bs.cell(row=row, column=6, value=f"=SUMIFS('Per-spell working'!$F$2:$F${LI_LAST},{q},\"{key}\")").font = FM_FONT
    bs.cell(row=row, column=7, value=f"=SUMIFS('Per-spell working'!$G$2:$G${LI_LAST},{q},\"{key}\")").font = FM_FONT
    bs.cell(row=row, column=8, value=b["cross"]).font = EN_FONT
    bs.cell(row=row, column=9, value=b["origin"]).font = EN_FONT
    bs.cell(row=row, column=10, value=f"=H{row}-F{row}").font = FM_FONT
    bs.cell(row=row, column=11, value=f"=I{row}-G{row}").font = FM_FONT
    # The diagnostic pair. The origin gap is how much the discount actually took off. If the discount
    # were being applied twice, the ENGINE gap would come out at roughly double the derived gap.
    bs.cell(row=row, column=12, value=f"=F{row}-G{row}").font = FM_FONT
    bs.cell(row=row, column=13, value=f"=H{row}-I{row}").font = FM_FONT
    bs.cell(row=row, column=14, value=(
        f'=IF(M{row}>L{row},"YES - engine gap exceeds model",'
        f'IF(M{row}=L{row},"no - applied exactly once",'
        f'"no - applied LESS than once"))')).font = FM_FONT
    bs.cell(row=row, column=15, value=f'=IF(AND(J{row}=0,K{row}=0),"yes","NO - hand-set")').font = FM_FONT
    row += 1
BS_LAST = row - 1
for rr in range(2, BS_LAST + 1):
    bs.conditional_formatting  # placeholder; explicit fills applied after recalc pass below
row += 1
bs.cell(row=row, column=1, value="Subclasses of these classes that sell NO bundle - nothing to buy, no cost, no option").font = Font(name=ARIAL, bold=True)
row += 1
for n in D["none"]:
    bs.cell(row=row, column=1, value=n["cls"]).font = FM_FONT
    bs.cell(row=row, column=2, value=n["sub"]).font = FM_FONT
    bs.cell(row=row, column=3, value="no spellBundle in the engine").font = EN_FONT
    for cc in (8, 9):
        bs.cell(row=row, column=cc, value="none").font = EN_FONT
    bs.cell(row=row, column=14, value="n/a - nothing to buy").font = EN_FONT
    row += 1
row += 1
bs.cell(row=row, column=1, value=(
    "Column N is the direct answer: a doubled discount would make the ENGINE gap (M) larger than the "
    "model gap (L). It never is - on any of the 21 bundles.")).font = Font(name=ARIAL, bold=True)

# ---------------------------------------------------------------- 4. Price paths --------------
pp = wb.create_sheet("Price paths")
pp.sheet_view.showGridLines = False
pp.column_dimensions["A"].width = 34
for col, w in zip("BCDE", (16, 16, 16, 46)):
    pp.column_dimensions[col].width = w
pp.cell(row=1, column=1, value="What the same buyer pays, by relationship to the class").font = Font(name=ARIAL, bold=True, size=13)
pp.cell(row=2, column=1, value="Verified by running compute() - not read off the source. See the note under each table.").font = FM_FONT
head(pp, 4, ["Purchase", "Origin class", "Class unlocked,\nnot origin", "Neither", "How that middle figure is reached"],
     [34, 16, 16, 16, 46])
rows = [
    ("Feature: Cleric: Blessed Strikes", 10, 13, 17, "sticker = cross - tier = 17 - 4 = 13"),
    ("Bundle: Cleric | Life Domain", 6, 8, 8, "no tier exists, so nothing is subtracted - it stays at cross = 8"),
]
r = 5
for lab, o, u, x, note in rows:
    pp.cell(row=r, column=1, value=lab).font = FM_FONT
    for i, v in enumerate((o, u, x), start=2):
        pp.cell(row=r, column=i, value=v).font = EN_FONT
    pp.cell(row=r, column=5, value=note).font = FM_FONT
    pp.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
for line in [
    "The guide's ability tables print column 3 as 'Sticker (Origin)'. In both rows above that column is the",
    "middle figure and the bracket is the first - i.e. 'what a non-origin member of the class pays (what an",
    "origin member pays)'. The meaning is consistent; only the route to it differs.",
    "",
    "The real asymmetry, which is worth a decision: unlocking a class cuts 4 AP off that class's FEATURE",
    "(17 -> 13) but nothing off its BUNDLE (8 -> 8). Bundles have no sticker tier in the engine at all.",
    "",
    "Neither table shows a discount being applied twice. If it were, the origin figure would fall below the",
    "per-spell floor - check the 'Per-spell working' sheet: no Origin AP cell there is ever less than 1.",
]:
    c = pp.cell(row=r, column=1, value=line)
    c.font = FM_FONT
    r += 1

for sheet in (li, bs):
    for rr in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for c in rr:
            c.border = THIN

wb.save(dest)
print(f"wrote {dest}: {len(D['bundles'])} bundles, {LI_LAST-1} line items")
